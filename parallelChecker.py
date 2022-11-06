from concurrent.futures import ThreadPoolExecutor
from threading import Thread

from selenium import webdriver
from gdev_api import units_from_profile
from openpyxl import load_workbook


def checkAndWriteSheet(profile_url, row_number):
    options = webdriver.EdgeOptions()
    options.headless = True
    driver = webdriver.Edge(options=options)
    print(f"", end="")

    try:
        units = units_from_profile(driver, profile_url)
    except Exception as e:
        print(f"[{row_number}]  [{profile_url}] gizli veya hatalı")
        ws[f"H{row_number}"] = "Hesap gizli veya link hatalı"
        return

    completed_unit_count = 0
    for unit, count in units.items():
        if count >= 3:
            completed_unit_count = completed_unit_count + 1
    print(f"[{row_number}]  [{profile_url}] | {completed_unit_count} ")
    ws[f"H{row_number}"] = f"{completed_unit_count}"
    ws[f"I{row_number}"] = f"{units['unit-1']}"
    ws[f"J{row_number}"] = f"{units['unit-2']}"
    ws[f"K{row_number}"] = f"{units['unit-3']}"
    ws[f"L{row_number}"] = f"{units['unit-4']}"
    ws[f"M{row_number}"] = f"{units['unit-5']}"
    driver.quit()


wb = load_workbook("Answers.xlsx")
ws = wb.active
devProfiles = ws[f"F2:F{ws.max_row}"]
ws[f"H1"] = "Tamamlanan Unit Sayısı"
ws[f"I1"] = "Unit1 Badge Sayısı"
ws[f"J1"] = "Unit2 Badge Sayısı"
ws[f"K1"] = "Unit3 Badge Sayısı"
ws[f"L1"] = "Unit4 Badge Sayısı"
ws[f"M1"] = "Unit5 Badge Sayısı"

with ThreadPoolExecutor(max_workers=20) as pool:
    rowNumber = 1
    for devProfile in devProfiles:
        rowNumber += 1
        profileUrl = str(devProfile[0].value).strip()
        if len(profileUrl) == 0:
            break
        if not (profileUrl.startswith("https://") or profileUrl.startswith("http://")):
            if profileUrl.startswith("g.dev"):
                profileUrl = "https://" + profileUrl
            else:
                profileUrl = "https://g.dev/" + profileUrl
        pool.submit(checkAndWriteSheet, profileUrl, rowNumber)

wb.save("CheckedList.xlsx")
