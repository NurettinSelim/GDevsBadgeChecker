from selenium import webdriver
from gdev_api import units_from_profile
from openpyxl import load_workbook

options = webdriver.EdgeOptions()
options.headless = True
driver = webdriver.Edge(options=options)

wb = load_workbook("Answers.xlsx")
ws = wb.active
devProfiles = ws["F"]
ws[f"H1"] = "Tamamlanan Unit Sayısı"
ws[f"I1"] = "Unit1 Badge Sayısı"
ws[f"J1"] = "Unit2 Badge Sayısı"
ws[f"K1"] = "Unit3 Badge Sayısı"
ws[f"L1"] = "Unit4 Badge Sayısı"
ws[f"M1"] = "Unit5 Badge Sayısı"

rowNumber = 0
for devProfile in devProfiles:
    rowNumber += 1

    profileUrl = str(devProfile.value).strip()
    if len(profileUrl) == 0:
        break
    if not (profileUrl.startswith("https://") or profileUrl.startswith("http://")):
        if profileUrl.startswith("g.dev"):
            profileUrl = "https://" + profileUrl
        else:
            profileUrl = "https://g.dev/" + profileUrl

    try:
        units = units_from_profile(driver, profileUrl)
    except Exception as e:
        print(f"[{profileUrl}] gizli veya hatalı")
        ws[f"H{rowNumber}"] = "Hesap gizli veya link hatalı"
        continue

    completed_unit_count = 0
    for unit, count in units.items():
        if count >= 3:
            completed_unit_count = completed_unit_count + 1
    print(f"{completed_unit_count} / {profileUrl}")
    ws[f"H{rowNumber}"] = f"{completed_unit_count}"
    ws[f"I{rowNumber}"] = f"{units['unit-1']}"
    ws[f"J{rowNumber}"] = f"{units['unit-2']}"
    ws[f"K{rowNumber}"] = f"{units['unit-3']}"
    ws[f"L{rowNumber}"] = f"{units['unit-4']}"
    ws[f"M{rowNumber}"] = f"{units['unit-5']}"

wb.save("CheckedList.xlsx")
